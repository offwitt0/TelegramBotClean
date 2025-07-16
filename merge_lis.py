import json
import pandas as pd
from openpyxl import load_workbook
from uuid import uuid4

def extract_hyperlinks(excel_path):
    """Extract hyperlinks from Unit Name column"""
    wb = load_workbook(excel_path)
    ws = wb.active
    hyperlinks = {}
    
    for row in range(2, ws.max_row + 1):  # Assuming header is row 1
        cell = ws.cell(row=row, column=1)  # Unit Name is column A
        if cell.hyperlink:
            hyperlinks[cell.value] = cell.hyperlink.target
    return hyperlinks

# Load existing listings
with open('listings.json', 'r', encoding='utf-8') as f:
    existing_listings = json.load(f)

# Extract hyperlinks from Excel
hyperlinks = extract_hyperlinks('AnQa.xlsx')

# Load Excel data
excel_data = pd.read_excel('AnQa.xlsx')

# Create new listings from Excel
new_listings = []
existing_names = {listing['name'].lower() for listing in existing_listings}

for _, row in excel_data.iterrows():
    # Skip if Unit Name is empty
    if pd.isna(row['Unit Name']):
        continue
    
    unit_name = str(row['Unit Name']).strip()
    listing_name = unit_name.split(':')[0].strip() if ':' in unit_name else unit_name
    
    # Skip if listing already exists (case-insensitive)
    if listing_name.lower() in existing_names:
        continue
    
    # Get URL from hyperlink if available
    url = hyperlinks.get(unit_name, "")
    
    # Set default values
    guests = int(row['Guests']) if pd.notna(row['Guests']) else 2
    bedrooms = int(row['Bedrooms #']) if pd.notna(row['Bedrooms #']) else max(1, guests // 2)
    bathrooms = float(row['Bathrooms #']) if pd.notna(row['Bathrooms #']) else 1
    
    # Create new listing
    new_listing = {
        'id': str(uuid4())[:8],
        'name': listing_name,
        'city_hint': str(row['Area']).strip() if pd.notna(row['Area']) else 'Cairo',
        'guests': guests,
        'bedrooms': bedrooms,
        'bathrooms': bathrooms,
        'rating': 4.5,
        'amenities': [
            "Kitchen",
            "Air conditioning",
            "Washing machine",
            "Wireless"
        ],
        'url': url
    }
    
    # Add conditional amenities
    if pd.notna(row['Parking']) and any(x in str(row['Parking']).lower() for x in ['available', 'yes', 'true']):
        new_listing['amenities'].append("Parking available")
    if pd.notna(row['Elevator']) and 'yes' in str(row['Elevator']).lower():
        new_listing['amenities'].append("Elevator")
    if pd.notna(row['Luggage']) and 'yes' in str(row['Luggage']).lower():
        new_listing['amenities'].append("Luggage storage")
    
    new_listings.append(new_listing)

# Merge with existing listings
merged_listings = existing_listings + new_listings

# Save the merged data
with open('listings.json', 'w', encoding='utf-8') as f:
    json.dump(merged_listings, f, indent=2, ensure_ascii=False)

print(f"Added {len(new_listings)} new listings with URLs. Total listings now: {len(merged_listings)}")